import json
import openpyxl
import math


ZONE_COUNT = 368


wb = openpyxl.load_workbook("1% Penetration Random Points.xlsx")
ws = wb.active
zones = ws["A"]
pointCounts = ws["B"]

pointCountInZone = dict()

for i in range(ZONE_COUNT + 1): 
    if zones[i].value is not None:
        pointCountInZone[zones[i].value] = pointCounts[i].value

# This is just to visualize dict in human readable format
# print(json.dumps(randomPointsByZoneIdDict, indent=4, sort_keys=True)) 







intersectionGeoJsonFile = "Intersection.geojson"
intersectionGeoJson = dict()
with open(intersectionGeoJsonFile) as _intersectionGeoJsonFile:
    intersectionGeoJson = json.load(_intersectionGeoJsonFile)


pointsInZone = dict()

# start iterating on features array in geojson and add them to pointsInZone, until pointsInZone reaches specified pointCount, for every zone
for point in intersectionGeoJson['features']:
    zoneId = point['properties']['Id_2']
    if pointsInZone.get(zoneId) is None:
        pointsInZone[zoneId] = []
    if len(pointsInZone[zoneId]) < pointCountInZone[zoneId]:
        pointsInZone[zoneId].append(point)


# extract feature array from every zone and make it a common all-zone feature array and dump output to json file cleanedZonePoints.json
features = []
for zoneId in pointsInZone:
    features += pointsInZone[zoneId]

with open('cleanedZonePoints.json', 'w') as cleanedZonePoints:
    json.dump({'features' : features}, cleanedZonePoints)






odMatrixWithPairedPoints = dict()

wb = openpyxl.load_workbook('2021 Stage OD Data.xlsx')
ws = wb.active

for fromZoneId in range(1, ZONE_COUNT + 1):
    for toZoneId in range(1, ZONE_COUNT + 1):
        frequencyCount = ws.cell(row=fromZoneId, column=toZoneId).value

        for pairId in range(math.floor(frequencyCount) - 1):

            if str((fromZoneId, toZoneId)) not in odMatrixWithPairedPoints:
                odMatrixWithPairedPoints[str((fromZoneId, toZoneId))] = []

            if fromZoneId in pointsInZone and toZoneId in pointsInZone:

                # if pairId >= len(pointsInZone[fromZoneId]):
                #     print(len(pointsInZone[fromZoneId]))
                #     print(fromZoneId)
                #     print(pairId)
                #     print('fromZoneId')

                # if pairId >= len(pointsInZone[toZoneId]):
                #     print(len(pointsInZone[toZoneId]))
                #     print(toZoneId)
                #     print('toZoneId')

                
                pair = [pointsInZone[fromZoneId][pairId], pointsInZone[toZoneId][pairId]]
                odMatrixWithPairedPoints[str((fromZoneId, toZoneId))].append(pair)



with open('odMatrixWithPairedPoints.json', 'w') as _odMatrixWithPairedPoints:
    json.dump(odMatrixWithPairedPoints, _odMatrixWithPairedPoints)
